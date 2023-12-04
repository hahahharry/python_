import math
from openpyxl import Workbook, load_workbook
import datetime
import numpy as np
import pandas as pd
import logging

logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

now = datetime.datetime.now()
morning = now.replace(hour=6, minute=30, second=0, microsecond=0)
day = now.replace(hour=10, minute=30, second=0, microsecond=0)
evening = now.replace(hour=17, minute=30, second=0, microsecond=0)

total_kcal, carbs_intake, protein_intake, fat_intake, kcal, carbs, proteins, fats = 0, 0, 0, 0, 0, 0, 0, 0

# System parameters
if morning < now < evening: # Daily ICR = 1/10
    icr = 1 / 10
elif now < day: # Morning ICR = 1/6
    icr = 1 / 6
else: # Evening ICR = 1/8
    icr = 1 / 8

target_bg = 4.5
corr_factor = 2.2

def check_recent_req_sugar(df, time_range_sugar):
    recent_required_sugar = []

    for index, row in df.iterrows():
        timestamp = row['Time']

        if timestamp >= time_range_sugar:
            recent_required_sugar.append(row['Req_sugar'])

    return sum(recent_required_sugar)

try:
    wb = load_workbook(filename='log.xlsx')
    sheet = wb['log']
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'log'
    sheet['A1'] = 'Time'
    sheet['B1'] = 'BG [mmol/L]'
    sheet['C1'] = 'Carbs [g]'
    sheet['D1'] = 'Bolus units'
    sheet['E1'] = 'Required Sugar'

db_worksheet = wb['Database']

while True:
    food_id = input("Enter food (type in 'End' to continue): ")
    logging.info("User entered food: %s", food_id)
    if food_id.lower() == "end":
        break
    serving_size_input = float(input("Enter serving size: "))

    for row in db_worksheet.iter_rows(min_row=2, min_col=1, values_only=True):

        if row[0] == food_id:
            serving_size = row[1]

            total_kcal += row[2] * serving_size_input / serving_size
            carbs_intake += row[3] * serving_size_input / serving_size
            protein_intake += row[4] * serving_size_input / serving_size
            fat_intake += row[5] * serving_size_input / serving_size
            break
    else:
        print("ID not found in the data!")
        print("Please input food manually")
        food_id = input("Food  name: ")
        serving_size = float(input("Serving size: "))
        actual_serving = float(input("Actual amount consumed: "))
        kcal += float(input("Total calories per serving: "))
        carbs += float(input("Total carbs per serving: "))
        proteins += float(input("Total protein per serving: "))
        fats += float(input("Total fat per serving: "))

        total_kcal = kcal * actual_serving / serving_size
        carbs_intake = carbs * actual_serving / serving_size
        protein_intake = proteins * actual_serving / serving_size
        fat_intake = fats * actual_serving / serving_size

        log_option = input('Do you want to save this food in the database? ("y" or "n"): ')

        if log_option == "y":
            new_row = [food_id, serving_size, kcal, carbs, proteins, fats]
            db_worksheet.append(new_row)
            print("Food added to the database!")

# User input
carbs = carbs_intake
curr_bg = float(input('Current BG [mmol/L]: '))
trend = int(input('BG trend (1 = sharp increase; 3 = plateau; 5 = sharp decrease): '))
activity = input('Physical activity ("y" or "n"): ')
sickness = input('Sickness ("y" or "n"): ')

if trend == 1:
    trend_factor = 5
elif trend == 2:
    trend_factor = 1.8
elif trend == 3:
    trend_factor = 0
elif trend == 4:
    trend_factor = -1.8
elif trend == 5:
    trend_factor = -5

if activity == "y":
    act_factor = 0.2
else:
    act_factor = 0

if sickness == "y":
    sick_factor = 0.2
else:
    sick_factor = 0

# Calculations
main_bolus = carbs * icr
correction_bolus = (curr_bg - target_bg) / corr_factor + trend_factor / corr_factor

if act_factor > 0 and correction_bolus < 0:
    adjustment_bolus = - (-act_factor + sick_factor) * (main_bolus + correction_bolus)
else:
    adjustment_bolus = (-act_factor + sick_factor) * (main_bolus + correction_bolus)

logging.info("Calculating main bolus: %f", main_bolus)
logging.info("Calculating correction bolus: %f", correction_bolus)
logging.info("Calculating adjustment bolus: %f", adjustment_bolus)

# Active insulin calculation
time = [0, 30, 60, 90, 120, 150, 180, 200]
insulin_activity = [0, 10, 10, 20, 20, 30, 7, 3]
cumulative_activity = np.cumsum(insulin_activity)

df = pd.read_excel('log.xlsx', names=['Time', 'BG', 'Carbs', 'Bolus', 'Protein', 'Fats', 'kcal', 'Req_sugar'], skiprows=[0])
df = df[['Time', 'Bolus', 'Carbs', 'Req_sugar']]
dia = 200  # Set the duration of insulin action to 200 minutes
values_list = []

if df['Bolus'].empty:
    most_recent_bolus = 0
    time1 = now
else:
    time_range = datetime.datetime.now() - datetime.timedelta(minutes=dia)

time_range_carbs = datetime.datetime.now() - datetime.timedelta(minutes=60)
active_insulin_dose, act_counter, recent_carbs_total = 0, 0, 0
recent_carbs = []

time_range_sugar = datetime.datetime.now() - datetime.timedelta(minutes=20)
recent_required_sugar_total = check_recent_req_sugar(df, time_range_sugar)

if recent_required_sugar_total > 0:
    main_bolus_timedelta = 0
    total_bolus = main_bolus_timedelta + correction_bolus + adjustment_bolus - active_insulin_dose
    required_sugar = recent_required_sugar_total - carbs_intake
else:
    main_bolus_timedelta = (carbs_intake * icr)
    total_bolus = main_bolus + main_bolus_timedelta + correction_bolus + adjustment_bolus - active_insulin_dose
    required_sugar = 0

for index, row in df.iterrows():
    timestamp = row[0]

    # Always calculate cumulative insulin and recent carbs
    time_diff = (timestamp - time_range).total_seconds() // 60
    cumulative_insulin_at_x = np.interp(time_diff, time, cumulative_activity)
    insulin_remaining = (cumulative_insulin_at_x / 100) * row[1]
    values_list.append(insulin_remaining)

    if timestamp >= time_range_carbs:
        recent_carbs.append(row[2])
        act_counter += 1

# Calculate active insulin and recent carbs after processing all rows
active_insulin_dose = sum(values_list)
recent_carbs_total = sum(recent_carbs)
logging.info("Active insulin calculated: %f", active_insulin_dose)

main_bolus_timedelta = (recent_carbs_total * icr)
total_bolus = main_bolus + main_bolus_timedelta + correction_bolus + adjustment_bolus - active_insulin_dose
logging.info("Total bolus calculated: %f",total_bolus)

last_row = sheet.max_row
while sheet.cell(row=last_row, column=1).value is None:
    last_row -= 1
last_row += 1

sheet.cell(row=last_row, column=1, value=now)
sheet.cell(row=last_row, column=2, value=curr_bg)
sheet.cell(row=last_row, column=3, value=carbs)
sheet.cell(row=last_row, column=5, value=protein_intake)
sheet.cell(row=last_row, column=6, value=fat_intake)
sheet.cell(row=last_row, column=7, value=total_kcal)

conf_bolus = 0
required_sugar = 0

if total_bolus < 0:
    required_sugar = math.ceil(abs(total_bolus / icr))
    print(f"Go munch something sweet!! You need {required_sugar}g of sugar!! ")
    total_bolus = 0
else:
    print(f"Total required bolus [units]: {math.ceil(total_bolus)} (main [{main_bolus:.1f}], correction [{correction_bolus + adjustment_bolus:.1f}])")
    print(f"There're {act_counter} active boluses with a total of {active_insulin_dose:.1f} units to be compensated.")
    conf_bolus = float(input('Confirm bolus (units): '))
    logging.info("User confirmed bolus: %.1f", conf_bolus)

sheet.cell(row=last_row, column=8, value=required_sugar)
sheet.cell(row=last_row, column=4, value=conf_bolus)
wb.close()
wb.save('log.xlsx')